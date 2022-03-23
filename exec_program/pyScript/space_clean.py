import json
import os
import xlwt
import xlrd
from xlutils.copy import copy
import re
import sys
import openpyxl


def get_num_from_str(str):
    num = re.findall(r'\d+',  str )
    if num!=[]:
        return num[0]
    else:
        return None


def remove_malloc(str):
    import  re
    patt_size = r'\d+'
    size_pattern = re.compile((patt_size))
    size_raw = size_pattern.findall(str)
    type = str.split("=")[1].split("malloc")[0].strip()[1:-2]

    if len(size_raw)>1:
        size = size_raw[1]
    elif len(size_raw) ==1:
        size = size_raw[0]
    else:
        size = "1"
    result = type+"*"+size
    return result


def extract_from_txt(txt_file):
    with open(txt_file, 'r') as f:
        data_raw = f.read()
    return data_raw


def space_clean(data_raw):
    """
    clean and extract important information
    """
    space_dict = {}
    data_set = data_raw.split("\n")
    for data_item in data_set:
        
        dir_name = 'data'
        data_detail = data_item.strip("(").strip(")").split(",")
        line = get_num_from_str(data_detail[0])
        if(line == None):
            continue
        space_line = data_detail[1]+":"+line
        func_name = data_detail[2]
        space_name = data_detail[3]+"@"+line
        if("ALLOCA" in data_detail or "malloc" in data_detail):
            space_size = data_detail[4]
        else:
            space_size = data_detail[4]+"*"+data_detail[5]
        space_type = "heap_space"
        loc = "heap_loc"        
        space_item = (space_line,func_name,space_name,space_type,loc,space_size)
        # file_raw = data_detail[1]
        
        # if (len(file_raw.split("/"))==8):
        #     dir_name = file_raw.split("/")[-2]  # s01,s02
        # else:
        #     dir_name ='s00'
        
        if dir_name in space_dict:
            space_dict[dir_name].append(space_item)
        else:
            space_dict[dir_name] = [space_item]
    
    return space_dict


def write_dict_to_excel(path,data,titles,sheet_name):
    """
    path : excel path to write
    data : a dict
    titles : first line of the excel
    sheet name : sheet name
    to write a dict to excel
    """
    book = openpyxl.Workbook()
    sheet = book.create_sheet(sheet_name)
    dir_name = "data"
    i = 1
    for title in titles:
        sheet.cell(1, i, title)
        i += 1
    if(data):
        for dir_name in data:

            for row in range(1, len(data[dir_name])):
                column = 1
                for j in data[dir_name][row - 1]:
                    sheet.cell(row + 1, column, str(j))
                    column += 1
    save_path = path+dir_name+"/"
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    book.save(save_path + "space.xls")



if __name__=="__main__":
    #src_path = "/home/iskindar/experiment/raw/uaf/"
    #dst_path = "/home/iskindar/experiment/clean/uaf/"
    src_path = sys.argv[1]
    dst_path = sys.argv[2]
    source_space = src_path + "space.txt"
    data_raw = extract_from_txt(source_space)
    space_dict = space_clean(data_raw)
    titles = ['line', 'func_name', 'space_name', 'space_type', 'loc','space_size']
    write_dict_to_excel(dst_path,space_dict,titles,"space")
    print("write space successfully",dst_path)
