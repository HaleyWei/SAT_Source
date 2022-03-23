import json
import os
import xlwt
import xlrd
from xlutils.copy import copy
import pickle
import sys
import openpyxl
def extact_data_from_json(json_file):

    with open(json_file, 'r') as f:
        data_raw = json.load(f)

    return data_raw


def ret_clean_new(data):
    ret_dict = {}
    dir_name = "data"
    for item in data:

        line = item["_2"][0] +":" +str(item["_1"])
        method_name = item["_3"]
        statement = item["_5"]
        argu = item['_4']
        if dir_name not in ret_dict:
            ret_dict[dir_name] =[(line,"ret",method_name,"ret",argu,"return","","","",statement)]
        else:
            ret_dict[dir_name].append((line,"ret",method_name,"ret",argu,"return","","","",statement))

    #print(ret_dict)
    return ret_dict



# def ret_clean(data_raw):
#     ret_list = {}

#     for item in data_raw:
#         if (item['_3']==[]):
#             continue
#         statement_num = len(item['_3'])
#         for i in range(0,statement_num):
#             method_name = item['_1']
#             file_raw = item['_2']
#             # if (len(file_raw.split("/"))==10):
#             #     dir_name = file_raw.split("/")[-2]  # s01,s02
#             # else:
#             #     dir_name ='s00'
#             dir_name = 's00'
#             statement = item['_3'][i]['_1']
#             line_raw = item['_3'][i]['_2']
#             argu_num = len(item['_3'][i]['_3'])
#             argu = ''
#             for j in range(0,argu_num):
#                 if (argu_num!=0):
#                     argu = item['_3'][i]['_3'][j]['code']

#             line = file_raw+":"+str(line_raw)
#             if dir_name not in ret_list:
#                 ret_list[dir_name] =[(line,"ret",method_name,"ret",argu,"return","","","",statement)]
#             else:
#                 ret_list[dir_name].append((line,"ret",method_name,"ret",argu,"return","","","",statement))

#     return ret_list

def func_call_clean_new(data):
    call_dict = {}
    dir_name = 'data'
    for item in data:
        argu_list = []  #to get argument_list
        argu_type_list = [] # to identify call_type
        
        line = item["_2"][0] +":"+ str(item["_5"])
        func_name = item["_3"]
        argu_full_list = item["_4"]
        func_called_name = item["_1"]
        statement = item["_6"]

         # call_type : normal_call, direct_call, indirect_call, libc_call, sys_call, ret
        sys_call = read_sys_call()
        libc_call = read_libc_call()
        call_type = ''

        for argu_item in argu_full_list:
            for argu_name in argu_item:
                
                if argu_item[argu_name] != []:
                    argu_list.append(argu_name)
                    argu_type_list.append(argu_item[argu_name][0])
        
        for argu_type in argu_type_list:
            if( len( match_bracket(argu_type) ) >= 2 ):
                call_type = 'indirect_call'
        if call_type =='':
            if func_called_name in sys_call:
                call_type = "sys_call"
            elif func_called_name in libc_call:
                call_type = "libc_call"
            elif "*" in func_called_name:
                call_type = "direct_call"
            else :
                call_type = "normal_call"
        
        if dir_name not in call_dict:
            call_dict[dir_name] = [(line, "call", func_name, call_type, argu_list, func_called_name, "", "", "", statement)]
        else:
            call_dict[dir_name].append((line, "call", func_name, call_type, argu_list, func_called_name, "", "", "", statement))
    #print(call_dict)
    return call_dict
        

# def func_call_clean(data_raw):
#     call_list = {}

#     for item in data_raw:
#         if (item['_3'] == []):
#             continue
#         statement_num = len(item['_3'])
#         func_name = item['_1']
#         file_raw = item['_2']

#         for i in range(0, statement_num):
#             argu_list = []  #to get argument_list
#             argu_type_list = [] # to identify call_type
#             func_called_name = item['_3'][i]['_1']
#             statement = item['_3'][i]['_2']
#             line_raw = item['_3'][i]['_3']
#             argu_num = len(item['_3'][i]['_4'])
#             line = file_raw + ":" + str(line_raw)
#             if (len(file_raw.split("/"))==8 ):
#                 dir_name = file_raw.split("/")[-2]  # s01,s02
#             else:
#                 dir_name ='s00'
#             #dir_name = 's00'
#             for j in range(0,argu_num):
#                 argu_list.append(item['_3'][i]['_4'][j]['code'])
#                 argu_type_list.append(item['_3'][i]['_4'][j]['typeFullName'])

#             # call_type : normal_call, direct_call, indirect_call, libc_call, sys_call, ret
#             sys_call = read_sys_call()
#             libc_call = read_libc_call()
#             call_type = ''
#             for argu_type in argu_type_list:
#                 if( len( match_bracket(argu_type) ) >= 2 ):
#                     call_type = 'indirect_call'
#             if call_type =='':
#                 if func_called_name in sys_call:
#                     call_type = "sys_call"
#                 elif func_called_name in libc_call:
#                     call_type = "libc_call"
#                 elif "*" in func_called_name:
#                     call_type = "direct_call"
#                 else :
#                     call_type = "normal_call"
#             if dir_name not in call_list:
#                 call_list[dir_name] = [(line, "call", func_name, call_type, argu_list, func_called_name, "", "", "", statement)]
#             else:
#                 call_list[dir_name].append((line, "call", func_name, call_type, argu_list, func_called_name, "", "", "", statement))

#     return call_list

def write_dict_to_excel(path,data,titles,sheet_name):
    """
    path : excel path to write
    data : a dict
    titles : first line of the excel
    sheet name : sheet name
    to write a dict to excel
    """
    book = openpyxl.Workbook()
    sheet = book.create_sheet(sheet_name)
    dir_name = "data"
    i = 1
    for title in titles:
        sheet.cell(1, i, title)
        i += 1
    if(data):
        for dir_name in data:

            for row in range(1, len(data[dir_name])):
                column = 1
                for j in data[dir_name][row - 1]:
                    sheet.cell(row + 1, column, str(j))
                    column += 1
    save_path = path+dir_name+"/"
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    book.save(save_path + "func_call.xls")


def read_libc_call():
    f = open('/home/weihaolai/ICKD_analysis/exec_program/pyScript/common_call_set/libc_call.pkl', 'rb')
    data = pickle.load(f)
    return data


def read_sys_call():
    f = open('/home/weihaolai/ICKD_analysis/exec_program/pyScript/common_call_set/sys_call.pkl', 'rb')
    data = pickle.load(f)
    return data


def match_bracket(str):
    import re
    patt = r'.*?(\(.*?\))'
    pattern = re.compile(patt)
    result = pattern.findall(str)
    return result

def main(source_file_ret,source_file_call,dst_path):
    result_dict = {}
    titles = ['line', 'state_type', 'func_name', 'call_type', 'func_arg', 'func_called_name', 'is_readorwrite',
              'safety', 'unsafe_reason', 'statement']
    ret_raw = extact_data_from_json(source_file_ret)
    ret_list = ret_clean_new(ret_raw)
    # write_list_to_excel("data/",ret_list,titles,"ret","ret.xls")
    call_raw = extact_data_from_json(source_file_call)
    call_list = func_call_clean_new(call_raw)
    ret_and_call = [ret_list,call_list]
    for dir_name in call_list:
        result_dict[dir_name] = []
        for list_item in ret_and_call:
            if(dir_name in list_item):
                result_dict[dir_name] += list_item[dir_name]
    write_dict_to_excel(dst_path, result_dict, titles, "func_call")
    # file_list = set()
    # for item in result_dict['s01']:
    #     file_name = item[0].split(":")[0]
    #     file_list.add(file_name)
    # print(len(file_list))

def test(source):
    data = extact_data_from_json(source)
    ret_clean_new(data)
if __name__=="__main__":
    
    # src_path = "/home/iskindar/experiment/raw/uaf/"
    # dst_path = "/home/iskindar/experiment/clean/uaf/"
    
    src_path = sys.argv[1]
    dst_path = sys.argv[2]
    
    source_file_ret = src_path+"ret.json"
    source_file_call = src_path+"call.json"
    
    main(source_file_ret,source_file_call,dst_path)
    print("write call successfully",dst_path)
