import json
import os
import xlwt
import xlrd
from xlutils.copy import copy
import sys
import openpyxl


def load_json(json_file):
    with open(json_file, 'r') as f:
        data = json.load(f)
    return data

def cond_clean_new(data):

    cond_dict = {}
    dir_name = 'data'
    for control_item in data:
        line = control_item[0][0] + ":" + str(control_item[1])
        method = control_item[2]
        argument_list = control_item[4]
        typ = control_item[3]
        start_line = control_item[1]
        if control_item[5] != []:
            end_line = max(control_item[5])
        else:
            end_line = ""
        statement = control_item[6]
        if dir_name not in cond_dict:
            cond_dict[dir_name] = [
                (line, "cond", method, argument_list, typ, start_line, end_line, statement)]
        else:
            cond_dict[dir_name].append(
                (line, "cond", method, argument_list, typ, start_line, end_line, statement))
    # print(cond_dict)
    return cond_dict
# def cond_clean(data):
#     cond_dict = {}

#     for item in data:
#         if item['_3']==[]:
#             continue
#         statement_num = len(item['_3'])
#         method = item['_1']
#         file_raw = item['_2']
#         for i in range(0,statement_num):

#             statement = item['_3'][i]
#             line_num = item['_4'][i]
#             if (len(file_raw.split("/"))==8):
#                 dir_name = file_raw.split("/")[-2]  # s01,s02
#             else:
#                 dir_name ='s00'
#             #dir_name = 's00'
#             argument_list_pre = item['_5']
#             argument_list = []
#             line = file_raw+":"+str(line_num)
#             # print(argument_list_pre)
#             # break
#             for argument_tuple in argument_list_pre:
#                 for argument in argument_tuple:
#                     if(argument_tuple[argument]==line_num):
#                         argument_list.append(argument)

#             if dir_name not in cond_dict:
#                 cond_dict[dir_name] = [(line,"cond",method,argument_list,statement)]
#             else:
#                 cond_dict[dir_name].append((line,"cond",method,argument_list,statement))

#     return cond_dict

# write data to excel


def write_dict_to_excel(path, data, titles, sheet_name):
    """
    path : excel path to write
    data : a dict
    titles : first line of the excel
    sheet name : sheet name
    to write a dict to excel
    """
    book = openpyxl.Workbook()
    sheet = book.create_sheet(sheet_name)
    dir_name = "data"
    i = 1
    for title in titles:
        sheet.cell(1, i, title)
        i += 1
    if(data):
        for dir_name in data:

            for row in range(1, len(data[dir_name])):
                column = 1
                for j in data[dir_name][row - 1]:
                    sheet.cell(row + 1, column, str(j))
                    column += 1
    save_path = path + dir_name + "/"
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    book.save(save_path + "cond.xls")


if __name__ == "__main__":
    titles = ['line', 'state_type', 'func_name', 'arg',
              'type', 'start_line', 'end_line', 'statement']
    # src_path = "/home/iskindar/experiment/raw/uaf/"
    # dst_path = "/home/iskindar/experiment/clean/uaf/"
    src_path = sys.argv[1]
    dst_path = sys.argv[2]
    source_cond = src_path + "cond.json"

    data_raw = load_json(source_cond)
    cond_clean_new(data_raw)

    cond_dict = cond_clean_new(data_raw)

    # print(cond_dict)
    # file_list= set()
    # for item in cond_dict['s01']:
    #     file_name = item[0].split(":")[0]
    #     file_list.add(file_name)
    # print(len(file_list))

    write_dict_to_excel(dst_path, cond_dict, titles, "cond")
    print("write cond successfully", dst_path)
