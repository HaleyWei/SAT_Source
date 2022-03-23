# coding:utf-8
from joern.all import JoernSteps
import db_search as search
import xlwt
import openpyxl
import sys


class define:
    def __init__(self, line, func_name, state_type, arg):
        self.line = line
        self.func_name = func_name
        self.state_type = state_type
        self.arg = arg


def get_def(db,src_path):
    list_id = search.getALLFuncNodeId_Name(db)
    file = "data/" + src_path + "/define.xls"
    # print file
    # book = xlwt.Workbook()
    book = openpyxl.Workbook()
    sheet = book.create_sheet('define')
    '''
    sheet.write(0, 0, 'line')
    sheet.write(0, 1, 'func_name')
    sheet.write(0, 2, 'state_type')
    sheet.write(0, 3, 'arg')
    '''
    sheet.cell(row=1, column=1).value = 'line'
    sheet.cell(row=1, column=2).value = 'func_name'
    sheet.cell(row=1, column=3).value = 'state_type'
    sheet.cell(row=1, column=4).value = 'arg'

    index_excle = 2
    define_excle = []

    for id_name in list_id:
        file_path = search.getFuncFile(db, id_name.id)
        print "********"
        print file_path + id_name.name
        node_list = search.getCFGNodes(db, id_name.id)
        var_list = search.get_fun_var(db, id_name.id)
        call_list = search.get_call_var(db, id_name.id)
        for node in node_list:
            if node.properties['location']:
                # print "--------"
                # print node.properties['code']
                def_var = search.get_def_node(db, node._id)
                if def_var:
                    node_line = ""
                    node_funcname = ""
                    node_state_type = ""
                    node_arg = ""
                    for d in def_var:
                        # 行号
                        node_line = str(
                            file_path) + ":" + str(node.properties['location'].split(':')[0])
                        # 函数名
                        node_funcname = id_name.name
                        # 语句类型
                        node_state_type = search.statement_Type(
                            search.get_type(db, node))
                        # 操作数
                        index_def = 0
                        d_code = d.properties['code']
                        if d_code not in call_list:
                            for var in var_list:
                                if var in d_code.split(' '):
                                    index_def = 1
                                    break
                            if index_def == 1:
                                d_var = d_code + '@' + id_name.name
                                node_arg += (d_var + " ")
                            else:
                                d_var = d_code + '@global'
                                node_arg += (d_var + " ")
                    # print node_line
                    # print node_funcname
                    # print node_state_type
                    # print node_arg
                    excle_item = define(node_line, node_funcname,
                                        node_state_type, node_arg)
                    define_excle.append(excle_item)

    for it in define_excle:
        sheet.cell(row=index_excle, column=1).value = it.line
        sheet.cell(row=index_excle, column=2).value = it.func_name
        sheet.cell(row=index_excle, column=3).value = it.state_type
        sheet.cell(row=index_excle, column=4).value = it.arg
        index_excle += 1
    book.save(file)


if __name__ == '__main__':
    src_path = sys.argv[1]
    j = JoernSteps()
    j.connectToDatabase()
    get_def(j,src_path)
