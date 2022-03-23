# # coding:utf-8
from joern.all import JoernSteps
import xlwt
import xlrd
import db_search as search
import open_table as table
import openpyxl
import sys
from py2neo.packages.httpstream import http
http.socket_timeout = 9999


class pointer:
    def __init__(self, line, name, safety, appear_line, unsafe_reason, unsafe_line, loc, pointer_type, point_to, key_data, scope, pointer_range):
        self.line = line
        self.name = name
        self.safety = safety
        self.appear_line = appear_line
        self.unsafe_reason = unsafe_reason
        self.unsafe_line = unsafe_line
        self.loc = loc
        self.pointer_type = pointer_type
        self.point_to = point_to
        self.key_data = key_data
        self.scope = scope
        self.pointer_range = pointer_range


class it_use:
    def __init__(self, target, result):
        self.target = target
        self.result = result


def get_pointers_node(db, func_id):
    list_pointers_node = []
    query_iddecl_str = 'queryNodeIndex("functionId:%s AND type:IdentifierDeclStatement")' % func_id
    results = db.runGremlinQuery(query_iddecl_str)
    if results != []:
        for re in results[::-1]:
            code = re.properties['code']
            if code.find(' = ') != -1:
                code = code.split(' = ')[0]
            if code.find('*') != -1:
                list_pointers_node.append(re)

    query_param_str = 'queryNodeIndex("functionId:%s AND type:Parameter")' % func_id
    results_param = db.runGremlinQuery(query_param_str)
    if results_param != []:
        for re in results_param[::-1]:
            code = re.properties['code']
            if code.find(' = ') != -1:
                code = code.split(' = ')[0]
            if code.find('*') != -1:
                list_pointers_node.append(re)

    return list_pointers_node


def get_arrays_node(db, func_id):
    list_arrays_node = []
    query_iddecl_str = "queryNodeIndex('functionId:%s AND type:IdentifierDeclStatement')" % func_id
    results = db.runGremlinQuery(query_iddecl_str)
    if results != []:
        for re in results:
            code = re.properties['code']
            if code.find(' = ') != -1:
                code = code.split(' = ')[0]

            if code.find(' [ ') != -1:
                list_arrays_node.append(re)

    query_param_str = "queryNodeIndex('functionId:%s AND type:Parameter')" % func_id
    results = db.runGremlinQuery(query_param_str)
    if results != []:
        for re in results:
            code = re.properties['code']
            if code.find(' = ') != -1:
                code = code.split(' = ')[0]

            if code.find(' [ ') != -1:
                list_arrays_node.append(re)

    return list_arrays_node


def get_all_use_bydefnode(db, node_id):
    query_str = "g.v(%d).in('USE')" % node_id
    results = db.runGremlinQuery(query_str)
    list_re = []
    for re in results:
        # print re
        if re.properties['type'] == 'Statement':
            continue
        else:
            list_re.append(re)

    return list_re


def get_all_def_bydefnode(db, node_id):
    query_str = "g.v(%d).in('DEF')" % node_id
    results = db.runGremlinQuery(query_str)
    list_re = []
    for re in results:
        # print re
        if re.properties['type'] == 'Statement':
            continue
        else:
            list_re.append(re)

    return list_re


def getDelete(db, func_id):
    query_str = 'queryNodeIndex("functionId:%s AND code:delete")' % func_id
    result = db.runGremlinQuery(query_str)
    list_delete = []
    for re in result:
        if re.properties['location'] != 'None':
            list_delete.append(re.properties['location'].split(':')[0])
    return list_delete


def get_mult(db, funcId):
    query_iddecl_str = 'queryNodeIndex("functionId:%d AND type:MultiplicativeExpression")' % funcId
    results = db.runGremlinQuery(query_iddecl_str)
    var = []
    if results != []:
        for re in results:
            code = re.properties['code']
            var.append(code)
    return var


def get_bit(db, funcId):
    query_iddecl_str = 'queryNodeIndex("functionId:%d AND type:BitAndExpression")' % funcId
    results = db.runGremlinQuery(query_iddecl_str)
    var = []
    if results != []:
        for re in results:
            code = re.properties['code']
            var.append(code)
    return var


def get_all_pointer_use(db, src_path):
    list_id = search.getALLFuncNodeId_Name(db)
    # 获取函数调用表
    func_call_list = table.get_func_call_table(src_path)
    # 获取条件调用表
    cond_list = table.get_cond_table(src_path)
    # 获取空间表
    space_list = table.get_space_table(src_path)

    file = 'data/' + src_path + '/pointer.xls'
    book = openpyxl.Workbook()
    sheet = book.create_sheet('pointer')
    sheet.cell(row=1, column=1).value = 'line'
    sheet.cell(row=1, column=2).value = 'name'
    sheet.cell(row=1, column=3).value = 'safety'
    sheet.cell(row=1, column=4).value = 'appear_line'
    sheet.cell(row=1, column=5).value = 'unsafe_reason'
    sheet.cell(row=1, column=6).value = 'unsafe_line'
    sheet.cell(row=1, column=7).value = 'loc'
    sheet.cell(row=1, column=8).value = 'pointer_type'
    sheet.cell(row=1, column=9).value = 'point_to'
    sheet.cell(row=1, column=10).value = 'key_data'
    sheet.cell(row=1, column=11).value = 'scope'
    sheet.cell(row=1, column=12).value = 'pointer_range'
    index_excle = 2
    pointer_excle = []

    for id_name in list_id:
        pointer_node = get_pointers_node(db, id_name.id)
        array_node = get_arrays_node(db, id_name.id)
        pointer_array_node = list(set(pointer_node + array_node))
        file_path = search.getFuncFile(db, id_name.id)
        pointer_use = search.getVarUse(db, id_name.id)
        # 获取函数中所有的变量
        var_list = search.get_fun_var(db, int(id_name.id))
        # 获取乘号的节点
        mult_state = get_mult(db, int(id_name.id))
        # 获取位与节点
        bit_state = get_bit(db, int(id_name.id))
        # delete关键字位置
        delete_loc = getDelete(db, id_name.id)
        print "------"
        print file_path + " " + id_name.name

        func_call = []
        for call in func_call_list:
            if call.file == file_path.split('/')[-1] and call.func_name == id_name.name:
                func_call.append(call)
        cond_state = []
        for cond in cond_list:
            if cond.file == file_path.split('/')[-1] and cond.func_name == id_name.name:
                cond_state.append(cond)
        space_state = []
        for space in space_list:
            if space.file == file_path.split('/')[-1] and space.func_name == id_name.name:
                space_state.append(space)
        pointer_var_use_list = []
        for use in pointer_use:
            target = use.properties['code'].strip()
            result = search.getLastNode(db, use._id)
            pointer_var_use_list.append(it_use(target, result))
        for node in pointer_array_node:
            # 指针定义节点
            pointer_def_node_list = search.get_def_node(db, node._id)
            for pNode_number in range(len(pointer_def_node_list)):
                if pNode_number != 10:
                    pointer_line = ""
                    pointer_name_func = ""
                    pointer_safety = "safe"
                    appear_line = ""
                    unsafe_reason = ""
                    unsafe_line = ""
                    pointer_loc = ""
                    pointer_type = ""
                    pointer_to = ""
                    key_data = "normal"
                    scope = "local"
                    pointer_range = ""
                    print "******"
                    # 指针名
                    # pointer_name = pointer_def_node.properties['code'].strip()
                    # point_name = "* " + pointer_name
                    # pointer_name_func = pointer_name + '@' + id_name.name
                    # 指针类型
                    pointer_type, pointer_name = search.getData_type(
                        db, node._id, pNode_number)
                    pointer_name_func = pointer_name + '@' + id_name.name
                    point_name = "* " + pointer_name
                    if "*" not in pointer_type and "[" not in pointer_type:
                        continue
                    # 指针定义位置
                    pointer_line = str(file_path) + ":" + \
                        str(node.properties['location'].split(':')[0])
                    # 获取指针出现的语句
                    list_use_node = []
                    for use in pointer_var_use_list:
                        if pointer_name == use.target:
                            if use.result != "null":
                                list_use_node.append(use.result)
                    list_use_node_set = list(set(list_use_node))
                    # 获取指针出现的行号
                    loc_list = []
                    for use in list_use_node_set:
                        loc_list.append(
                            int(use.properties['location'].split(':')[0]))
                    loc_list.sort()
                    for loc_item in loc_list:
                        appear_line += str(loc_item) + " "
                    # 判断指针的存储位置
                    exist_static, loc_static = search.getStaticNode(
                        db, node._id)
                    if exist_static and loc_static == node.properties['location'].split(':')[0]:
                        pointer_loc = "static_data_loc"
                    elif "new" in str(node.properties['code']) or "malloc" in str(node.properties['code']):
                        pointer_loc = "heap_loc"
                    else:
                        pointer_loc = "stack_loc"
                    # 判断指针是否是非安全指针
                    for use in list_use_node_set:
                        code = use.properties['code']
                        code_loc = use.properties['location'].split(':')[0]
                        # 非安全类型是free_to_ptr
                        if "free (" in code:
                            unsafe_line += (code_loc + " ")
                            pointer_safety = "unsafe"
                            unsafe_reason += "free_to_ptr "
                        if code_loc in delete_loc:
                            unsafe_line += (code_loc + " ")
                            pointer_safety = "unsafe"
                            unsafe_reason += "free_to_ptr "
                        # 非安全类型是num_to_ptr
                        # 如果是数组，数组内部有变量，看作有缓冲区溢出特征
                        if(code.find('%s [' % pointer_name) != -1):
                            code_sc = code.split('%s [' % pointer_name)[1]
                            end = code_sc.find(']')
                            array_var_list = code_sc[:end].split(" ")
                            for var in array_var_list:
                                if var in var_list:
                                    unsafe_line += (code_loc + " ")
                                    pointer_safety = "unsafe"
                                    unsafe_reason += "buf_to_ptr "
                        elif(code.find(' = ') != -1):
                            a_code = code.split(' = ')[1]
                            if (a_code.find('+') != -1 or a_code.find('-') != -1 or a_code.find('/') != -1 or a_code.find('%') != -1 or a_code.find('++') != -1 or a_code.find('--') != -1 or a_code.find('|') != -1 or a_code.find('^') != -1 or a_code.find('~') != -1 or a_code.find('<<') != -1 or a_code.find('>>') != -1):
                                if (pointer_name in code.split(' = ')[0].split(' ')) and (point_name not in str(code.split(' = ')[0])):
                                    unsafe_line += (code_loc + " ")
                                    pointer_safety = "unsafe"
                                    unsafe_reason += "num_to_ptr "
                            elif(a_code.find('*') != -1):
                                if (pointer_name in code.split(' = ')[0].split(' ')) and (point_name not in str(code.split(' = ')[0])):
                                    for m in mult_state:
                                        if m in a_code:
                                            unsafe_line += (code_loc + " ")
                                            pointer_safety = "unsafe"
                                            unsafe_reason += "num_to_ptr "
                            elif(a_code.find('&') != -1):
                                if (pointer_name in code.split(' = ')[0].split(' ')) and (point_name not in str(code.split(' = ')[0])):
                                    for b in bit_state:
                                        if b in a_code:
                                            unsafe_line += (code_loc + " ")
                                            pointer_safety = "unsafe"
                                            unsafe_reason += "num_to_ptr "
                        elif(code.find('+=') != -1 or code.find('-=') != -1 or code.find('*=') != -1 or code.find('/=') != -1 or code.find('%=') != -1 or code.find('<<=') != -1 or code.find('>>=') != -1 or code.find('&=') != -1 or code.find('^=') != -1 or code.find('|=') != -1):
                            if (pointer_name in code.split('=')[0].split(' ')) and (point_name not in str(code.split('=')[0])):
                                unsafe_line += (code_loc + " ")
                                pointer_safety = "unsafe"
                                unsafe_reason += "num_to_ptr "
                        elif(code.find('++') != -1 or code.find('--') != -1):
                            if point_name not in str(code):
                                unsafe_line += (code_loc + " ")
                                pointer_safety = "unsafe"
                                unsafe_reason += "num_to_ptr "
                        # 判断指向的关键数据与漏洞特征情况
                        for call in func_call:
                            if call.line == code_loc:
                                # 判断是否是关键数据
                                if (pointer_name in call.arg) and (call.call_type == "libc_call" or call.call_type == "sys_call"):
                                    key_data = "key_non_control_data"
                                # 判断是否有缓冲区溢出特征
                                if call.call_name in table.bof_func and call.arg != [] and pointer_name == call.arg[0]:
                                    unsafe_line += (code_loc + " ")
                                    pointer_safety = "unsafe"
                                    unsafe_reason += "buf_to_ptr "

                        for cond in cond_state:
                            if cond.line == code_loc:
                                if pointer_name in cond.arg:
                                    key_data = "key_non_control_data"
                        # 判断指针范围
                        if pointer_type.find('[') != -1 and pointer_range == "":
                            num_long = pointer_type.split(
                                '[')[-1].split(']')[0].strip()
                            # print num_long
                            pointer_range += str(pointer_type.split('[')[0].replace(
                                '*', '').strip()) + "*(" + str(num_long) + ")"
                        else:
                            for space in space_state:
                                if space.line == code_loc and (pointer_name in code.split(" = ")[0].split(' ')):
                                    pointer_range = space.space_size

                    excle_item = pointer(pointer_line, pointer_name_func, pointer_safety, appear_line, unsafe_reason,
                                         unsafe_line, pointer_loc, pointer_type, pointer_to, key_data, scope, pointer_range)
                    pointer_excle.append(excle_item)

    for it in pointer_excle:
        sheet.cell(row=index_excle, column=1).value = it.line
        sheet.cell(row=index_excle, column=2).value = it.name
        sheet.cell(row=index_excle, column=3).value = it.safety
        sheet.cell(row=index_excle, column=4).value = it.appear_line
        sheet.cell(row=index_excle, column=5).value = it.unsafe_reason
        sheet.cell(row=index_excle, column=6).value = it.unsafe_line
        sheet.cell(row=index_excle, column=7).value = it.loc
        sheet.cell(row=index_excle, column=8).value = it.pointer_type
        sheet.cell(row=index_excle, column=9).value = it.point_to
        sheet.cell(row=index_excle, column=10).value = it.key_data
        sheet.cell(row=index_excle, column=11).value = it.scope
        sheet.cell(row=index_excle, column=12).value = it.pointer_range

        index_excle += 1
    book.save(file)

    # print index


if __name__ == '__main__':
    j = JoernSteps()
    j.connectToDatabase()
    src_path = sys.argv[1]
    get_all_pointer_use(j, src_path)
