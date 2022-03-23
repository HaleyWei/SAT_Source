# coding:utf-8
from joern.all import JoernSteps
import db_search as search
import xlwt
import openpyxl
import sys

class func:
    def __init__(self, f_name, f_type, f_arg, s_line, f_line):
        self.name = f_name
        self.type = f_type
        self.arg = f_arg
        self.start = s_line
        self.finish = f_line


def get_funcType(db, func_id):
    func_type = ""
    query_str = "queryNodeIndex('functionId:%s AND type:ReturnType')" % func_id
    results = db.runGremlinQuery(query_str)
    for re in results:
        func_type = re.properties['code']
    return func_type


def get_funcParameter(db, func_id):
    var = ""
    query_iddecl_pra = 'queryNodeIndex("functionId:%d AND type:Parameter")' % func_id
    results = db.runGremlinQuery(query_iddecl_pra)
    if results != []:
        for re in results:
            def_node = search.get_def_node(db, re._id)
            for db_node in def_node:
                var += (db_node.properties['code'] + " ")
    return var


def get_funcLocation(db, func_id):
    query = 'queryNodeIndex("functionId:%d AND isCFGNode:True")' % func_id
    results = db.runGremlinQuery(query)
    location_list = []
    for re in results:
        if re.properties['location']:
            location_list.append(int(re.properties['location'].split(':')[0]))
    location_list = sorted(location_list)
    return location_list


def get_func(db,src_path):
    list_func = search.getALLFuncNode(db)
    file = 'data/' + src_path + '/func.xls'
    # book = xlwt.Workbook()
    book = openpyxl.Workbook()
    sheet = book.create_sheet('func')
    sheet.cell(row=1, column=1).value = 'func_name'
    sheet.cell(row=1, column=2).value = 'func_type'
    sheet.cell(row=1, column=3).value = 'func_arg'
    sheet.cell(row=1, column=4).value = 'start_line'
    sheet.cell(row=1, column=5).value = 'finish_line'
    '''
    sheet.write(0, 0, 'func_name')
    sheet.write(0, 1, 'func_type')
    sheet.write(0, 2, 'func_arg')
    sheet.write(0, 3, 'start_line')
    sheet.write(0, 4, 'finish_line')
    '''
    index_excle = 2
    func_excle = []

    for func_node in list_func:
        func_name = ""
        func_type = ""
        func_arg = ""
        start_line = ""
        finish_line = ""
        # 函数名与路径
        file_path = search.getFuncFile(db, func_node._id)
        print "--------"
        func_name = str(file_path) + ":" + str(func_node.properties['name'])
        # 函数类型
        func_type = get_funcType(db, func_node._id)
        # 函数参数
        func_arg = get_funcParameter(db, func_node._id)
        # 函数终止与起始位置
        location = get_funcLocation(db, func_node._id)
        if location:
            start_line = str(location[0])
            finish_line = str(location[-1])
        print "func:" + file_path
        # print func_name
        # print func_type
        # print func_arg
        # print start_line
        # print finish_line
        excle_item = func(func_name, func_type, func_arg,
                          start_line, finish_line)
        func_excle.append(excle_item)

    for it in func_excle:
        '''
        sheet.write(index_excle, 0, it.name)
        sheet.write(index_excle, 1, it.type)
        sheet.write(index_excle, 2, it.arg)
        sheet.write(index_excle, 3, it.start)
        sheet.write(index_excle, 4, it.finish)
        '''
        sheet.cell(row=index_excle, column=1).value = it.name
        sheet.cell(row=index_excle, column=2).value = it.type
        sheet.cell(row=index_excle, column=3).value = it.arg
        sheet.cell(row=index_excle, column=4).value = it.start
        sheet.cell(row=index_excle, column=5).value = it.finish
        index_excle += 1
    book.save(file)


if __name__ == '__main__':
    j = JoernSteps()
    j.connectToDatabase()
    src_path = sys.argv[1] 
    get_func(j,src_path)
