# # coding:utf-8
from joern.all import JoernSteps
import xlwt
import db_search as search
import open_table as table
import openpyxl
import sys
from py2neo.packages.httpstream import http
http.socket_timeout = 9999


class data_var:
    def __init__(self, line, name, data_type, data_content, key_data, loc, source, status, scope, data_use):
        self.line = line
        self.name = name
        self.data_type = data_type
        self.data_content = data_content
        self.key_data = key_data
        self.loc = loc
        self.source = source
        self.status = status
        self.scope = scope
        self.data_use = data_use


class it_use:
    def __init__(self, target, result):
        self.target = target
        self.result = result


def get_memcpy_arg(db, node):
    result = search.getNextNode(db, node._id)
    arg_list = []
    if result != []:
        for re in result:
            if re.properties['type'] == "Argument":
                arg_list.append(re.properties['code'])
            else:
                arg = get_memcpy_arg(db, re)
                if arg:
                    arg_list += arg
    if arg_list:
        return arg_list


def get_all_data_use(db, src_path):
    list_id = search.getALLFuncNodeId_Name(db)
    # 获取函数调用表
    func_call_list = table.get_func_call_table(src_path)
    # 获取条件调用表
    cond_list = table.get_cond_table(src_path)

    # 用户输入函数
    input_func = ['scanf', 'gets', 'getline',
                  'get', 'getchar', 'getche', 'getch', 'fgets'
                  'fgetc', 'read', 'fopen', 'fread', 'fscanf']

    file = 'data/' + src_path + '/data.xls'
    # book = xlwt.Workbook()
    book = openpyxl.Workbook()
    sheet = book.create_sheet('data')
    '''
    sheet.write(0, 0, 'line')
    sheet.write(0, 1, 'name')
    sheet.write(0, 2, 'data_type')
    sheet.write(0, 3, 'data_content')
    sheet.write(0, 4, 'key_data')
    sheet.write(0, 5, 'loc')
    sheet.write(0, 6, 'source')
    sheet.write(0, 7, 'status')
    sheet.write(0, 8, 'scope')
    sheet.write(0, 9, 'data_use')
    '''
    sheet.cell(row=1, column=1).value = 'line'
    sheet.cell(row=1, column=2).value = 'name'
    sheet.cell(row=1, column=3).value = 'data_type'
    sheet.cell(row=1, column=4).value = 'data_content'
    sheet.cell(row=1, column=5).value = 'key_data'
    sheet.cell(row=1, column=6).value = 'loc'
    sheet.cell(row=1, column=7).value = 'source'
    sheet.cell(row=1, column=8).value = 'status'
    sheet.cell(row=1, column=9).value = 'scope'
    sheet.cell(row=1, column=10).value = 'data_use'
    index_excle = 2
    data_excle = []

    for id_name in list_id:

        data_node = search.get_all_data(db, id_name.id)
        file_path = search.getFuncFile(db, id_name.id)
        data_var_use = search.getVarUse(db, id_name.id)
        print "var:" + file_path
        func_call = []
        for call in func_call_list:
            if call.file == file_path.split('/')[-1] and call.func_name == id_name.name:
                func_call.append(call)
        cond_state = []
        for cond in cond_list:
            if cond.file == file_path.split('/')[-1] and cond.func_name == id_name.name:
                cond_state.append(cond)
        # 收集数据使用及其源节点
        data_var_use_list = []
        for use in data_var_use:
            target = use.properties['code'].strip()
            result = search.getLastNode(db, use._id)
            data_var_use_list.append(it_use(target, result))
        for data in data_node:
            data_def_node_list = search.get_def_node(db, data._id)
            for dNode_number in range(len(data_def_node_list)):
                if dNode_number != 10:                    
                    data_line = ""
                    data_name_func = ""
                    data_type = ""
                    data_content = ""
                    key_data = "normal"
                    data_loc = ""
                    data_source = "input"
                    data_status = "dynamic_data"
                    data_scope = "local"
                    data_use = ""
                    # 数据名和类型
                    data_type, data_name = search.getData_type(
                        db, data._id, dNode_number)
                    # 数据行号
                    data_line = str(file_path) + ":" + \
                        str(data.properties['location'].split(':')[0])
                    # data_name = (search.get_def_node(db, data._id))[-1].properties['code'].strip()
                    data_name_func = data_name + '@' + id_name.name
                    # 数据的存储位置
                    exist_static, loc_static = search.getStaticNode(
                        db, data._id)
                    if exist_static and loc_static == data.properties['location'].split(':')[0]:
                        data_loc = "static_data_loc"
                        data_status = "static_data"
                    elif "new" in str(data.properties['code']) or "malloc" in str(data.properties['code']):
                        data_loc = "heap_loc"
                    else:
                        data_loc = "stack_loc"
                    # 获取数据出现的语句
                    list_use_node = []
                    for use in data_var_use_list:
                        if data_name == use.target:
                            if use.result != "null":
                                list_use_node.append(use.result)
                    '''
                    for use in data_var_use:
                        if data_name == use.properties['code'].strip():
                            the_use = search.getLastNode(db, use._id)
                            if the_use != "null":
                                list_use_node.append(the_use)
                    '''
                    list_use_node_set = list(set(list_use_node))
                    for use in list_use_node_set:
                        code = use.properties['code']
                        code_loc = use.properties['location'].split(':')[0]
                        # 数据内容
                        if "memcpy" in code.split(' '):
                            arg_list = get_memcpy_arg(db, use)
                            if arg_list[0] == data_name:
                                data_content = arg_list[1]
                        # 是否是关键数据
                        for call in func_call:
                            if call.line == code_loc:
                                if (data_name in call.arg) and (call.call_type == "libc_call" or call.call_type == "sys_call"):
                                    key_data = "key_non_control_data"
                        for cond in cond_state:
                            if cond.line == code_loc:
                                if data_name in cond.arg:
                                    key_data = "key_non_control_data"
                        # 是否是外部输入-1
                        for put in input_func:
                            if put in code.split(' '):
                                data_source = 'output'
                        # 数据用途
                        for call in func_call:
                            if call.line == code_loc:
                                if (data_name in call.arg) and (call.call_type == "normal_call"):
                                    data_use += "func_arg "
                                if (data_name in call.arg) and (call.call_type == "libc_call"):
                                    data_use += "libc_arg "
                                if (data_name in call.arg) and (call.call_type == "sys_call"):
                                    data_use += "sys_arg "
                        for cond in cond_state:
                            if cond.line == code_loc:
                                if data_name in cond.arg:
                                    data_use += "cond_arg "
                    if data_use == "":
                        data_use = "other"
                    # 是否是外部输入-2
                    if data.properties['type'] == 'Parameter':
                        data_source = 'output'

                    # print data_line
                    # print data_name_func
                    # print data_type
                    # print data_content
                    # print key_data
                    # print data_loc
                    # print data_source
                    # print data_status
                    # print data_scope
                    # print data_use
                    excle_item = data_var(data_line, data_name_func, data_type, data_content,
                                          key_data, data_loc, data_source, data_status, data_scope, data_use)
                    data_excle.append(excle_item)
    for it in data_excle:
        '''
        sheet.write(index_excle, 0, it.line)
        sheet.write(index_excle, 1, it.name)
        sheet.write(index_excle, 2, it.data_type)
        sheet.write(index_excle, 3, it.data_content)
        sheet.write(index_excle, 4, it.key_data)
        sheet.write(index_excle, 5, it.loc)
        sheet.write(index_excle, 6, it.source)
        sheet.write(index_excle, 7, it.status)
        sheet.write(index_excle, 8, it.scope)
        sheet.write(index_excle, 9, it.data_use)
        '''
        sheet.cell(row=index_excle, column=1).value = it.line
        sheet.cell(row=index_excle, column=2).value = it.name
        sheet.cell(row=index_excle, column=3).value = it.data_type
        sheet.cell(row=index_excle, column=4).value = it.data_content
        sheet.cell(row=index_excle, column=5).value = it.key_data
        sheet.cell(row=index_excle, column=6).value = it.loc
        sheet.cell(row=index_excle, column=7).value = it.source
        sheet.cell(row=index_excle, column=8).value = it.status
        sheet.cell(row=index_excle, column=9).value = it.scope
        sheet.cell(row=index_excle, column=10).value = it.data_use
        index_excle += 1
    book.save(file)


if __name__ == '__main__':
    j = JoernSteps()
    j.connectToDatabase()
    src_path = sys.argv[1]
    get_all_data_use(j, src_path)
