# coding:utf-8
from joern.all import JoernSteps
import db_search as search
import open_table as table
import xlwt
import openpyxl
import sys


class op:
    def __init__(self, line, func_name, state_type, src, dst, is_readorwrite, safety, unsafe_reason):
        self.line = line
        self.func_name = func_name
        self.state_type = state_type
        self.src = src
        self.dst = dst
        self.is_readorwrite = is_readorwrite
        self.safety = safety
        self.unsafe_reason = unsafe_reason


def get_op_node(db, func_id):
    op_list = []
    query_str_1 = 'queryNodeIndex("functionId:%d AND type:AssignmentExpr")' % func_id
    results_1 = db.runGremlinQuery(query_str_1)
    for re in results_1:
        op_list.append(re)
    query_str_2 = 'queryNodeIndex("functionId:%d AND type:IncDecOp")' % func_id
    results_2 = db.runGremlinQuery(query_str_2)
    for re in results_2:
        op_list.append(re)
    return op_list


def get_op(db,src_path):
    list_id = search.getALLFuncNodeId_Name(db)

    file = 'data/' + src_path + '/op.xls'
    # book = xlwt.Workbook()
    book = openpyxl.Workbook()
    sheet = book.create_sheet('op')
    '''
    sheet.write(0, 0, 'line')
    sheet.write(0, 1, 'func_name')
    sheet.write(0, 2, 'state_type')
    sheet.write(0, 3, 'src')
    sheet.write(0, 4, 'dst')
    sheet.write(0, 5, 'is_readorwrite')
    sheet.write(0, 6, 'safety')
    sheet.write(0, 7, 'unsafe_reason')
    '''
    sheet.cell(row=1, column=1).value = 'line'
    sheet.cell(row=1, column=2).value = 'func_name'
    sheet.cell(row=1, column=3).value = 'state_type'
    sheet.cell(row=1, column=4).value = 'src'
    sheet.cell(row=1, column=5).value = 'dst'
    sheet.cell(row=1, column=6).value = 'is_readorwrite'
    sheet.cell(row=1, column=7).value = 'safety'
    sheet.cell(row=1, column=8).value = 'unsafe_reason'
    index_excle = 2
    op_excle = []

    for id_name in list_id:
        file_path = search.getFuncFile(db, id_name.id)
        print "********"
        print file_path + id_name.name
        op_node = get_op_node(db, id_name.id)
        var_list = search.get_fun_var(db, id_name.id)
        call_list = search.get_call_var(db, id_name.id)
        for node in op_node:
            node_line = ""
            node_funcname = ""
            node_state_type = ""
            node_src = ""
            node_dst = ""
            node_readorwrite = ""
            node_safety = ""
            node_reason = ""
            # 行号
            # print node.properties['code']
            if node.properties['isCFGNode'] == "True":
                init_node = node
            else:
                init_node = search.getLastNode(db, node._id)
            node_line = str(file_path) + ":" + \
                str(init_node.properties['location'].split(':')[0])
            # 所在函数名
            node_funcname = id_name.name
            # 语句类型
            node_state_type = "op"
            # 源操作数和目的操作数
            def_var = search.get_def_node(db, init_node._id)
            use_var = search.get_use_node(db, init_node._id)
            def_var_name = []
            use_var_name = []
            for d in def_var:
                index_def = 0
                d_code = d.properties['code']
                if d_code not in call_list:
                    def_var_name.append(d_code)
                    for var in var_list:
                        if var in d_code.split(' '):
                            index_def = 1
                            break
                    if index_def == 1:
                        d_var = d_code + '@' + id_name.name
                        node_dst += (d_var + " ")
                    else:
                        d_var = d_code + '@global'
                        node_dst += (d_var + " ")
            # 输出源操作数，判断函数中的变量是否存在于操作数中
            for u in use_var:
                index_use = 0
                u_code = u.properties['code']
                if u_code not in call_list:
                    use_var_name.append(u_code)
                    for var in var_list:
                        if var in u_code.split(' '):
                            index_use = 1
                            break
                    if index_use == 1:
                        u_var = u_code + '@' + id_name.name
                        node_src += (u_var + " ")
                    else:
                        u_var = u_code + '@global'
                        node_src += (u_var + " ")
            # 判断是否进行内存读写，等号左边变量为内存写，等号右边有变量为内存读
            store = 0
            load = 0
            if def_var:
                for d in def_var:
                    if d.properties['code'] not in call_list:
                        store = 1
                        break
            if use_var:
                for u in use_var:
                    if u.properties['code'] not in call_list:
                        load = 1
                        break
            if store == 1 and load == 0:
                node_readorwrite = "store"
            elif store == 0 and load == 1:
                node_readorwrite = "load"
            elif store == 1 and load == 1:
                node_readorwrite = "store load"
            else:
                node_readorwrite = "no"
            # print node_line
            # print node_funcname
            # print node_state_type
            # print node_src
            # print node_dst
            # print node_readorwrite
            # print node_safety
            # print node_reason
            excle_item = op(node_line, node_funcname, node_state_type,
                            node_src, node_dst, node_readorwrite, node_safety, node_reason)
            op_excle.append(excle_item)

    for it in op_excle:
        '''
        sheet.write(index_excle, 0, it.line)
        sheet.write(index_excle, 1, it.func_name)
        sheet.write(index_excle, 2, it.state_type)
        sheet.write(index_excle, 3, it.src)
        sheet.write(index_excle, 4, it.dst)
        sheet.write(index_excle, 5, it.is_readorwrite)
        sheet.write(index_excle, 6, it.safety)
        sheet.write(index_excle, 7, it.unsafe_reason)
        '''
        sheet.cell(row=index_excle, column=1).value = it.line
        sheet.cell(row=index_excle, column=2).value = it.func_name
        sheet.cell(row=index_excle, column=3).value = it.state_type
        sheet.cell(row=index_excle, column=4).value = it.src
        sheet.cell(row=index_excle, column=5).value = it.dst
        sheet.cell(row=index_excle, column=6).value = it.is_readorwrite
        sheet.cell(row=index_excle, column=7).value = it.safety
        sheet.cell(row=index_excle, column=8).value = it.unsafe_reason
        index_excle += 1
    book.save(file)


if __name__ == '__main__':
    j = JoernSteps()
    j.connectToDatabase()
    src_path = sys.argv[1]
    get_op(j,src_path)
