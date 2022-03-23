# coding:utf-8
from joern.all import JoernSteps
import db_search as search
import xlwt
import openpyxl
import sys


class def_use:
    def __init__(self, line, state_type, src_arg, dst_arg, src, dst):
        self.line = line
        self.state_type = state_type
        self.src_arg = src_arg
        self.dst_arg = dst_arg
        self.src = src
        self.dst = dst


def getDef_Use(db, src_path):
    list_id = search.getALLFuncNodeId_Name(db)
    file = 'data/' + src_path + '/def_use.xls'
    # book = xlwt.Workbook()
    book = openpyxl.Workbook()
    sheet = book.create_sheet('def_use')
    sheet.cell(row=1, column=1).value = 'line'
    sheet.cell(row=1, column=2).value = 'state_type'
    sheet.cell(row=1, column=3).value = 'src_arg'
    sheet.cell(row=1, column=4).value = 'dst_arg'
    sheet.cell(row=1, column=5).value = 'src'
    sheet.cell(row=1, column=6).value = 'dst'
    index_excle = 2
    def_use_excle = []

    for id_name in list_id:
        file_path = search.getFuncFile(db, id_name.id)
        print "********"
        print file_path + ' ' + id_name.name
        # node_list = search.getCFGNodes(db, id_name.id)
        var_list = search.get_fun_var(db, id_name.id)
        call_list = search.get_call_var(db, id_name.id)
        # 源、目的语句的列表
        src_dst_list = []
        # 根据函数获得边
        node_list = set()
        node_list_2 = set()    # 去重
        edge_list = search.getDDGEdges(db, id_name.id)
        for edge in edge_list:
            # print edge
            e = search.item(edge.start_node, edge.end_node)
            src_dst_list.append(e)
            node_list.add(edge.start_node)
            node_list.add(edge.end_node)
        for node in node_list:
            if node.properties['location'] and node not in node_list_2:
                node_list_2.add(node)
                node_line = ""
                node_state_type = ""
                node_src_arg = ""
                node_dst_arg = ""
                node_src = ""
                node_dst = ""
                # print "--------"
                # print node
                # print node.properties['code']
                # 行号获取*********************
                node_line = str(file_path) + ":" + \
                    str(search.getLocation(node))
                # 类型获取*********************
                node_state_type = search.statement_Type(
                    search.get_type(db, node))
                # node_type_list.append(node_type)
                def_var = search.get_def_node(db, node._id)
                use_var = search.get_use_node(db, node._id)
                # 输出目的操作数，判断函数中的变量是否存在于操作数中
                for d in def_var:
                    index_def = 0
                    d_code = d.properties['code']
                    if d_code not in call_list:
                        for var in var_list:
                            if var in d_code.split(' '):
                                index_def = 1
                                break
                        # 目的操作数获取*********************
                        if index_def == 1:
                            d_var = d_code + '@' + id_name.name
                            node_dst_arg += (d_var + " ")
                        else:
                            d_var = d_code + '@global'
                            node_dst_arg += (d_var + " ")
                # 输出源操作数，判断函数中的变量是否存在于操作数中
                for u in use_var:
                    index_use = 0
                    u_code = u.properties['code']
                    if u_code not in call_list:
                        for var in var_list:
                            if var in u_code.split(' '):
                                index_use = 1
                                break
                        # 源操作数获取***********************
                        if index_use == 1:
                            u_var = u_code + '@' + id_name.name
                            node_src_arg += (u_var + " ")
                        else:
                            u_var = u_code + '@global'
                            node_src_arg += (u_var + " ")
                # 获得源和目的语句的行号
                src = []
                dst = []
                for s in src_dst_list:
                    if node == s.src:
                        dst.append(s.dst)
                    if node == s.dst:
                        src.append(s.src)
                # 根据语句来源获得操作数来源
                node_src_line = []
                node_dst_line = []
                if src:
                    for s in src:
                        # +2020.8.23
                        src_line = str(search.getLocation(s))
                        node_src_line.append(src_line)
                # 根据语句目的获得操作数目的
                if dst:
                    for d in dst:
                        # +2020.8.23
                        dst_line = str(search.getLocation(d))
                        node_dst_line.append(dst_line)
                # +2020.8.23
                node_src_line = list(set(node_src_line))
                node_dst_line = list(set(node_dst_line))
                for line in node_src_line:
                    node_src += line + " "
                for line in node_dst_line:
                    node_dst += line + " "

                # print node_line
                # print node_state_type
                # print node_src_arg
                # print node_dst_arg
                # print node_src
                # print node_dst
                print "--------"
                excle_item = def_use(
                    node_line, node_state_type, node_src_arg, node_dst_arg, node_src, node_dst)
                def_use_excle.append(excle_item)

    for it in def_use_excle:
        sheet.cell(row=index_excle, column=1).value = it.line
        sheet.cell(row=index_excle, column=2).value = it.state_type
        sheet.cell(row=index_excle, column=3).value = it.src_arg
        sheet.cell(row=index_excle, column=4).value = it.dst_arg
        sheet.cell(row=index_excle, column=5).value = it.src
        sheet.cell(row=index_excle, column=6).value = it.dst
        index_excle += 1
    book.save(file)


if __name__ == '__main__':
    j = JoernSteps()
    j.connectToDatabase()
    src_path = sys.argv[1]
    getDef_Use(j, src_path)
